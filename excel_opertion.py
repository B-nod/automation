import os

import openpyxl
import glob
test_result_location = r'C:\QA deer walk\code_for_automation\facebook_test_case.xlsx'


# def clear_result():
#     files = glob.glob('/Users/Shared/Private/Python/sqa_training/sqa_20/test_result/*')
#     for f in files:
#         os.remove(f)
#     print("All file has been removed")


def write_header():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Test Result")
    ws.cell(1, 1, "SN")
    ws.cell(1, 2, "Test Summary")
    ws.cell(1, 3, "Result")
    ws.cell(1, 4, "Remarks")
    wb.save(test_result_location)


def write_result(sn, test_summary, result, remarks):
    wb = openpyxl.load_workbook(test_result_location)
    ws = wb["Test Result"]
    row = int(sn)+1
    ws.cell(row, 1, sn)
    ws.cell(row, 2, test_summary)
    ws.cell(row, 3, result)
    ws.cell(row, 4, str(remarks))
    wb.save(test_result_location)


